import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

from config import MASTER_CSV


def _safe_cell(value) -> str | int | float | None:
    if value is None:
        return None
    s = str(value)
    if s[:1] in ("=", "+", "-", "@"):
        return "'" + s
    return value


def export_audit_to_excel(df: pd.DataFrame, filename: str = "Audit_Report.xlsx", output_dir: str | None = None) -> str:
    output_dir = output_dir or os.environ.get("EXPORT_DIR", "/data/exports")
    os.makedirs(output_dir, exist_ok=True)

    safe_name = os.path.basename(filename)
    if not safe_name.lower().endswith(".xlsx"):
        safe_name += ".xlsx"
    path = os.path.join(output_dir, safe_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Data"
    ws["A1"] = "Belfield Pharmacy Internal Audit Report"
    ws["A1"].font = Font(bold=True, size=14)

    for col_idx, column in enumerate(df.columns, start=1):
        ws.cell(row=3, column=col_idx, value=column)

    for r_idx, row in enumerate(df.values.tolist(), start=4):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_safe_cell(value))

    wb.save(path)
    os.chmod(path, 0o600)
    return path


def load_master_dataframe() -> pd.DataFrame:
    if not os.path.exists(MASTER_CSV):
        return pd.DataFrame(columns=["Date", "Task", "Patient_ID", "Staff_Assigned", "Status"])
    return pd.read_csv(MASTER_CSV)
